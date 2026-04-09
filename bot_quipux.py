import asyncio
import os
import re
import sys
from datetime import datetime, date
from playwright.async_api import async_playwright

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠ openpyxl no instalado. Ejecuta: pip install openpyxl")
    print("  El Excel no se generará, pero las descargas continuarán.\n")


# ─────────────────────────────────────────────────────────
# Generar archivo Excel a partir de datos de documentos
# ─────────────────────────────────────────────────────────
def generar_excel(all_docs_data, xlsx_path, sheet_name="Documentos"):
    if not HAS_OPENPYXL or not all_docs_data:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="006394", end_color="006394", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Calibri", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    even_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

    headers = ["No.", "De", "Asunto", "Fecha Documento",
               "Número Documento", "No. Referencia",
               "Usuario Anterior", "Carpeta Descarga", "Anexos"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, doc_data in enumerate(all_docs_data, 2):
        values = [
            doc_data["No"], doc_data["De"], doc_data["Asunto"],
            doc_data["Fecha Documento"], doc_data["Número Documento"],
            doc_data["No. Referencia"], doc_data["Usuario Anterior"],
            doc_data["Carpeta"], doc_data["Anexos"],
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = even_fill

    col_widths = [6, 35, 50, 18, 25, 18, 30, 30, 8]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(xlsx_path)
    print(f"\n  📊 Excel generado: {xlsx_path}")


# ─────────────────────────────────────────────────────────
# Parsear fecha de documento Quipux a objeto date
# ─────────────────────────────────────────────────────────
def parsear_fecha(fecha_str):
    """Intenta parsear una fecha en varios formatos comunes de Quipux.
    Retorna un objeto date o None si no se puede parsear."""
    fecha_str = fecha_str.strip()
    formatos = [
        "%d/%m/%Y",      # 06/04/2026
        "%Y-%m-%d",      # 2026-04-06
        "%d-%m-%Y",      # 06-04-2026
        "%d/%m/%Y %H:%M",  # 06/04/2026 14:30
        "%Y-%m-%d %H:%M",  # 2026-04-06 14:30
        "%d-%m-%Y %H:%M",  # 06-04-2026 14:30
        "%d/%m/%Y %H:%M:%S",  # 06/04/2026 14:30:00
        "%Y-%m-%d %H:%M:%S",  # 2026-04-06 14:30:00
    ]
    for fmt in formatos:
        try:
            return datetime.strptime(fecha_str, fmt).date()
        except ValueError:
            continue
    return None


# ─────────────────────────────────────────────────────────
# Función para navegar a una bandeja específica
# ─────────────────────────────────────────────────────────
QUIPUX_BASE = "https://quipux.espe.edu.ec"

async def navegar_bandeja(target_page, bandeja_nombre, carpeta_codigo, pagina=1, es_paginacion=False):
    """Navega a la bandeja indicada haciendo CLICK en el enlace del menú.
    Si no puede navegar automáticamente, pide al usuario que haga clic manual.
    Retorna el mainFrame o None."""

    # ── Función auxiliar: obtener mainFrame ──
    def get_main_frame():
        for f in target_page.frames:
            if f.name == "mainFrame":
                return f
        for f in target_page.frames:
            if f.url and "cuerpo.php" in f.url:
                return f
        return None

    # ── Función auxiliar: verificar bandeja cargada ──
    async def verificar_bandeja(mf):
        """Retorna el nombre de la bandeja que realmente se muestra."""
        try:
            info = await mf.evaluate("""
                () => {
                    const body = document.body.innerText || '';
                    // Buscar "Bandeja: Enviados" o "Bandeja: Recibidos"
                    const match = body.match(/Bandeja:\\s*(\\S+)/);
                    return match ? match[1] : 'desconocido';
                }
            """)
            return info
        except Exception:
            return "error"

    # ── Si es paginación, usar paginador dentro de la bandeja actual ──
    if es_paginacion and pagina > 1:
        mf = get_main_frame()
        if mf:
            try:
                await mf.evaluate(
                    f"paginador_reload_div('adodb_next_page={pagina}')"
                )
                await asyncio.sleep(3)
            except Exception:
                pass
        return get_main_frame()

    # ── Diagnóstico: listar frames ──
    print(f"       Frames disponibles:")
    for i, f in enumerate(target_page.frames):
        try:
            fu = f.url[:80] if f.url else "N/A"
        except Exception:
            fu = "error"
        print(f"         [{i}] name='{f.name}', url={fu}")

    navegacion_ok = False

    # ── Método 1: Buscar enlace llamarListado en menú (comillas simples y dobles) ──
    print(f"       Buscando enlace de menú para '{bandeja_nombre}'...")
    for frame in target_page.frames:
        if frame.name == "mainFrame":
            continue
        try:
            link_text = await frame.evaluate(f"""
                () => {{
                    // Buscar con comillas simples y dobles
                    const targets = [
                        "llamarListado('{bandeja_nombre}'",
                        'llamarListado("{bandeja_nombre}"',
                        "llamarListado(\\'{bandeja_nombre}\\'",
                    ];
                    const links = document.querySelectorAll('a');
                    document.querySelectorAll('[data-bot-target]').forEach(
                        el => el.removeAttribute('data-bot-target')
                    );
                    for (const link of links) {{
                        const href = (link.getAttribute('href') || '');
                        const onclick = (link.getAttribute('onclick') || '');
                        const combined = href + ' ' + onclick;
                        for (const target of targets) {{
                            if (combined.includes(target)) {{
                                link.setAttribute('data-bot-target', 'true');
                                return link.innerText.trim();
                            }}
                        }}
                    }}
                    return false;
                }}
            """)
            if link_text:
                await frame.click('a[data-bot-target="true"]')
                await frame.evaluate(
                    "document.querySelector('[data-bot-target]')"
                    "?.removeAttribute('data-bot-target')"
                )
                navegacion_ok = True
                print(f"       ✓ Click en '{link_text}' (frame: {frame.name or '?'})")
                break
        except Exception:
            continue

    # ── Método 2: Buscar enlace cuyo texto COMIENCE con el nombre ──
    if not navegacion_ok:
        print(f"       Buscando enlace por texto...")
        for frame in target_page.frames:
            if frame.name == "mainFrame":
                continue
            try:
                found = await frame.evaluate(f"""
                    () => {{
                        const links = document.querySelectorAll('a');
                        document.querySelectorAll('[data-bot-target]').forEach(
                            el => el.removeAttribute('data-bot-target')
                        );
                        for (const link of links) {{
                            const text = link.innerText.trim();
                            // Coincide "Enviados" o "Enviados (3)" pero NO "No Enviados"
                            if (text === '{bandeja_nombre}' ||
                                text.match(/^{bandeja_nombre}\\s*\\(\\d/)) {{
                                link.setAttribute('data-bot-target', 'true');
                                return text;
                            }}
                        }}
                        return false;
                    }}
                """)
                if found:
                    await frame.click('a[data-bot-target="true"]')
                    await frame.evaluate(
                        "document.querySelector('[data-bot-target]')"
                        "?.removeAttribute('data-bot-target')"
                    )
                    navegacion_ok = True
                    print(f"       ✓ Click en '{found}' (frame: {frame.name or '?'})")
                    break
            except Exception:
                continue

    # Esperar a que el contenido se actualice
    if navegacion_ok:
        await asyncio.sleep(5)

    # ── Verificar qué bandeja se cargó realmente ──
    mf = get_main_frame()
    if mf:
        bandeja_real = await verificar_bandeja(mf)
        print(f"       Bandeja cargada: {bandeja_real}")
        if bandeja_nombre != "Carpetas Virtuales" and bandeja_nombre.lower() not in bandeja_real.lower():
            print(f"       ⚠ La bandeja cargada ({bandeja_real}) NO coincide "
                  f"con la solicitada ({bandeja_nombre})")
            navegacion_ok = False  # Forzar fallback manual

    # ── FALLBACK MANUAL: pedir al usuario que haga clic ──
    if not navegacion_ok:
        print(f"\n  ╔══════════════════════════════════════════════════════╗")
        print(f"  ║  ⚠ No se pudo navegar automáticamente a {bandeja_nombre:<12}║")
        print(f"  ║  Por favor, haz clic en '{bandeja_nombre}' en el menú    ║")
        print(f"  ║  lateral izquierdo del navegador.                    ║")
        print(f"  ╚══════════════════════════════════════════════════════╝")
        input(f"\n  >>> Presiona ENTER cuando hayas seleccionado '{bandeja_nombre}' en el navegador... ")
        await asyncio.sleep(2)
        mf = get_main_frame()
        if mf:
            bandeja_real = await verificar_bandeja(mf)
            print(f"       Bandeja cargada: {bandeja_real}")

    return get_main_frame()


# ─────────────────────────────────────────────────────────
# Descargar todos los documentos de la bandeja actual
# ─────────────────────────────────────────────────────────
async def descargar_bandeja(target_page, main_frame, context, download_dir, bandeja_nombre, carpeta_codigo="8", fecha_desde=None, fecha_hasta=None):
    """Descarga todos los documentos (con paginación) de la bandeja activa.
    Si fecha_desde y fecha_hasta (date) se proporcionan, sólo descarga documentos
    cuya fecha esté dentro del rango [fecha_desde, fecha_hasta] (inclusive).
    carpeta_codigo: '2' para Recibidos, '8' para Enviados (usado en paginación)."""

    # Extraer nombre simple de bandeja (sin " — usuario")
    bandeja_simple = bandeja_nombre.split(' — ')[0] if ' — ' in bandeja_nombre else bandeja_nombre

    os.makedirs(download_dir, exist_ok=True)

    # Esperar a que se cargue el contenido
    try:
        await main_frame.wait_for_selector("tr.listado1, tr.listado2", timeout=15000)
    except Exception:
        print("       Esperando carga AJAX...")
        await asyncio.sleep(5)

    # Verificar URL y contenido del frame cargado
    try:
        frame_url = main_frame.url
        print(f"       Frame URL: {frame_url}")
        # Verificar qué bandeja se muestra realmente
        bandeja_detectada = await main_frame.evaluate("""
            () => {
                const body = document.body.innerText || '';
                // Buscar indicadores de qué bandeja estamos viendo
                const url = location.href || '';
                const nomcarpeta = url.match(/nomcarpeta=([^&]+)/);
                return {
                    url_carpeta: nomcarpeta ? nomcarpeta[1] : 'desconocido',
                    url_completa: location.href
                };
            }
        """)
        print(f"       Bandeja en URL: {bandeja_detectada.get('url_carpeta', '?')}")
    except Exception as e:
        print(f"       ⚠ Error verificando bandeja: {e}")

    # Detectar total de páginas
    total_pages = await main_frame.evaluate("""
        () => {
            const text = document.body.innerText;
            const match = text.match(/Página\\s+(\\d+)\\/(\\d+)/);
            if (match) return parseInt(match[2]);
            return 1;
        }
    """)
    print(f"       Páginas encontradas: {total_pages}")
    if fecha_desde and fecha_hasta:
        print(f"       📅 Rango de fechas: {fecha_desde.strftime('%d/%m/%Y')} → {fecha_hasta.strftime('%d/%m/%Y')}")
    else:
        print(f"       📅 Descargando TODOS los documentos")

    descargados = 0
    errores = 0
    total_docs = 0
    omitidos_fecha = 0
    all_docs_data = []

    for current_page in range(1, total_pages + 1):
        print(f"\n{'=' * 60}")
        print(f"  📄 PÁGINA {current_page} / {total_pages} — {bandeja_nombre}")
        print(f"{'=' * 60}")

        # Navegar a la página si no es la primera
        if current_page > 1:
            print(f"  Navegando a página {current_page}...")
            main_frame = await navegar_bandeja(
                target_page, bandeja_simple, carpeta_codigo,
                pagina=current_page, es_paginacion=True
            )
            if not main_frame:
                print(f"  ⚠ Error: No se pudo navegar a página {current_page}")
                break
            try:
                await main_frame.wait_for_selector(
                    "tr.listado1, tr.listado2", timeout=15000
                )
            except Exception:
                await asyncio.sleep(3)

        # Extraer documentos de la página actual
        docs = await main_frame.evaluate("""
            () => {
                const docs = [];
                // Quitar restricción de clases para atrapar los resultados de Carpetas Virtuales también
                const rows = document.querySelectorAll('tr');
                const isSearchLayout = location.href.includes('lista_documentos_buscar.php');
                
                rows.forEach((row, idx) => {
                    const style = window.getComputedStyle(row);
                    if (style.display === 'none' || style.visibility === 'hidden') return;
                    if (row.offsetParent === null && style.position !== 'fixed') return;

                    const cells = row.querySelectorAll('td');
                    // Ignorar la fila si no tiene las celdas mínimas
                    if (cells.length < 5) return;
                    
                    let radicado = '', textrad = '', carpeta = '';
                    const allLinks = row.querySelectorAll('a');
                    for (const link of allLinks) {
                        const href = link.getAttribute('href') || '';
                        const onclick = link.getAttribute('onclick') || '';
                        const text = onclick + ' ' + href;
                        
                        // Permisivo: Buscar cualquier función típica de Quipux que maneje apertura de documentos
                        // ej: mostrar_documento('1234','5432','8') o ver_datos_documento('1234','')
                        const match = text.match(/(?:mostrar|ver|abrir)_.*?[a-zA-Z0-9]*documento.*?[\\(][^'"\\d]*['"](\\d+)['"](?:[\\s,]+['"]([^'"]*)['"])?(?:[\\s,]+['"]([^'"]*)['"])?/i);
                        if (match && match[1]) {
                            radicado = match[1];
                            textrad = match[2] || match[1];
                            carpeta = match[3] || '';
                            break;
                        }
                    }
                    
                    // Extraer basándose en el layout de la tabla
                    let asunto = '', remitente = '', fecha = '', numDoc = '';
                    let noReferencia = '', usuarioAnterior = '';
                    
                    if (isSearchLayout && cells.length >= 7) {
                        // Layout de Carpetas Virtuales: De(0), Para(1), Asunto(2), Fecha(3), Num(4), Usuario(5), Estado(6)
                        remitente = (cells[0]?.innerText || '').trim();
                        // El Para está en cells[1] si deseas sumarlo después
                        asunto = (cells[2]?.innerText || '').trim();
                        fecha = (cells[3]?.innerText || '').trim().replace(/\\s*\\(GMT.*?\\)/i, '');
                        numDoc = (cells[4]?.innerText || '').trim();
                    } else if (cells.length >= 8) {
                        // Layout tradicional bandejas
                        remitente = (cells[5]?.innerText || '').trim();
                        asunto = (cells[6]?.innerText || '').trim();
                        fecha = (cells[7]?.innerText || '').trim();
                        numDoc = (cells[8]?.innerText || '').trim();
                        
                        if (cells.length >= 10) noReferencia = (cells[9]?.innerText || '').trim();
                        if (cells.length >= 11) usuarioAnterior = (cells[10]?.innerText || '').trim();
                    }

                    // Solo incluir filas que tengan radicado descubiertos por el script nativo 
                    if (radicado && (numDoc || asunto || remitente)) {
                        docs.push({
                            index: idx, radicado, textrad, carpeta,
                            asunto, remitente, fecha, numDoc,
                            noReferencia, usuarioAnterior
                        });
                    }
                });
                return docs;
            }
        """)

        print(f"  Documentos en esta página: {len(docs)}")

        if not docs:
            debug_html = await main_frame.evaluate("""
                () => {
                    const row = document.querySelector('tr.listado1, tr.listado2') || document.querySelectorAll('tr')[5];
                    return row ? row.outerHTML : 'No se encontraron filas <tr> para debug.';
                }
            """)
            print("\n  [DEBUG] Ningún documento fue extraído de esta página.")
            print("  Si crees que esto es un error, por favor revisa el código HTML de una fila de ejemplo:")
            print("  " + "-"*40)
            print(f"  {debug_html[:800]}...")
            print("  " + "-"*40)
            print("  Sin documentos, saltando...")
            continue

        # ── LOG DETALLADO: mostrar todos los documentos extraídos ──
        print(f"\n  {'No.':<5} {'Carpeta':<15} {'Num. Documento':<30} {'Remitente':<25} {'Fecha':<12}")
        print(f"  {'-'*5} {'-'*15} {'-'*30} {'-'*25} {'-'*12}")
        for i, d in enumerate(docs, 1):
            print(f"  {i:<5} {d.get('carpeta','?'):<15} {d.get('numDoc','?')[:28]:<30} "
                  f"{d.get('remitente','?')[:23]:<25} {d.get('fecha','?'):<12}")
        print()

        # Variable para controlar si debemos detener por fecha
        detener_por_fecha = False

        for doc in docs:
            total_docs += 1
            radicado = doc["radicado"]
            textrad = doc["textrad"]
            carpeta = doc["carpeta"]
            asunto = doc["asunto"][:80]
            numDoc = doc["numDoc"]
            noReferencia = doc.get("noReferencia", "")
            usuarioAnterior = doc.get("usuarioAnterior", "")

            # ── Filtro por rango de fechas ──
            if fecha_desde and fecha_hasta:
                fecha_doc = parsear_fecha(doc.get("fecha", ""))
                if fecha_doc:
                    if fecha_doc < fecha_desde:
                        omitidos_fecha += 1
                        print(f"  ─── [{total_docs}] {numDoc} ───")
                        print(f"  ⏭ Omitido (fecha {doc.get('fecha', '?')} anterior a {fecha_desde.strftime('%d/%m/%Y')})")
                        print()
                        continue
                    if fecha_doc > fecha_hasta:
                        omitidos_fecha += 1
                        print(f"  ─── [{total_docs}] {numDoc} ───")
                        print(f"  ⏭ Omitido (fecha {doc.get('fecha', '?')} posterior a {fecha_hasta.strftime('%d/%m/%Y')})")
                        print()
                        continue

            safe_num = re.sub(r'[^\w\-]', '_', numDoc or textrad)
            doc_folder = f"DOC_{total_docs:03d}_{safe_num}"
            doc_dir = os.path.join(download_dir, doc_folder)
            anexos_dir = os.path.join(doc_dir, "anexos")
            os.makedirs(doc_dir, exist_ok=True)

            print(f"  ─── [{total_docs}] {numDoc} ───")
            print(f"  Carpeta: {carpeta}")
            print(f"  Asunto: {asunto}")
            print(f"  De: {doc.get('remitente', '?')}")

            try:
                doc_url = (
                    f"https://quipux.espe.edu.ec/verradicado.php"
                    f"?verrad={radicado}&textrad={textrad}"
                    f"&carpeta={carpeta}&menu_ver_tmp=3&tipo_ventana=popup"
                )
                print(f"  URL: {doc_url}")
                doc_page = await context.new_page()
                await doc_page.goto(doc_url, wait_until="domcontentloaded", timeout=60000)
                await asyncio.sleep(3)

                # ---- DOCUMENTO PRINCIPAL ----
                doc_downloaded = False

                ver_doc_selectors = [
                    'a:has-text("Ver Documento")',
                    'a:has-text("Ver documento")',
                    'a:has-text("ver documento")',
                    'a[href*="ver_documento"]',
                    'a[href*="anexos_descargar_archivo"]',
                ]
                for sel in ver_doc_selectors:
                    try:
                        ver_link = await doc_page.query_selector(sel)
                        if ver_link:
                            try:
                                async with doc_page.expect_download(timeout=15000) as dl:
                                    await ver_link.click()
                                download = await dl.value
                                fname = download.suggested_filename or f"documento_{total_docs}.pdf"
                                await download.save_as(os.path.join(doc_dir, fname))
                                print(f"  ✓ Documento: {fname}")
                                doc_downloaded = True
                                descargados += 1
                            except Exception:
                                pass
                        if doc_downloaded:
                            break
                    except Exception:
                        continue

                if not doc_downloaded:
                    try:
                        iframe_src = await doc_page.evaluate("""
                            () => {
                                const ifr = document.getElementById('ifr_descargar_archivo');
                                return ifr ? ifr.src : null;
                            }
                        """)
                        if iframe_src and 'radi_nume' in iframe_src:
                            dl_page = await context.new_page()
                            try:
                                async with dl_page.expect_download(timeout=15000) as dl:
                                    await dl_page.goto(iframe_src)
                                download = await dl.value
                                fname = download.suggested_filename or f"documento_{total_docs}.pdf"
                                await download.save_as(os.path.join(doc_dir, fname))
                                print(f"  ✓ Documento (iframe): {fname}")
                                doc_downloaded = True
                                descargados += 1
                            except Exception:
                                pass
                            finally:
                                await dl_page.close()
                    except Exception:
                        pass

                if not doc_downloaded:
                    try:
                        async with doc_page.expect_download(timeout=15000) as dl:
                            await doc_page.evaluate("""
                                () => {
                                    if (typeof vista_previa === 'function') vista_previa();
                                    else if (typeof fjs_radicado_descargar_archivo === 'function')
                                        fjs_radicado_descargar_archivo(radi_nume, '', 0, 'download');
                                }
                            """)
                        download = await dl.value
                        fname = download.suggested_filename or f"documento_{total_docs}.pdf"
                        await download.save_as(os.path.join(doc_dir, fname))
                        print(f"  ✓ Documento (JS): {fname}")
                        doc_downloaded = True
                        descargados += 1
                    except Exception:
                        pass

                if not doc_downloaded:
                    print(f"  ⚠ No se pudo descargar el documento principal.")

                # ---- ANEXOS ----
                print(f"  Buscando anexos...")
                try:
                    anexos_url = (
                        f"https://quipux.espe.edu.ec/verradicado.php"
                        f"?carpeta={carpeta}&verrad={radicado}"
                        f"&textrad={textrad}&estadisticas=0"
                        f"&verPDF=1&irVerRad=1&tipo_ventana=popup&menu_ver=2"
                    )
                    await doc_page.goto(anexos_url, wait_until="domcontentloaded", timeout=60000)
                    await asyncio.sleep(2)

                    div_status = 'loading'
                    for _ in range(12):
                        div_status = await doc_page.evaluate("""
                            () => {
                                const div = document.getElementById('div_anexos_lista_archivos');
                                if (!div) return 'no_div';
                                const html = div.innerHTML.trim();
                                if (html === '') return 'empty';
                                if (html.includes('no tiene archivos anexos')) return 'sin_anexos';
                                if (html.includes('anexos_descargar_archivo')) return 'tiene_anexos';
                                if (html.length > 50) return 'loaded';
                                return 'loading';
                            }
                        """)
                        if div_status in ('sin_anexos', 'tiene_anexos', 'loaded', 'no_div'):
                            break
                        await asyncio.sleep(0.5)

                    if div_status == 'sin_anexos':
                        print(f"  (Sin anexos)")
                    elif div_status == 'no_div':
                        print(f"  ⚠ div_anexos no encontrado")
                    else:
                        anexo_data = await doc_page.evaluate("""
                            () => {
                                const div = document.getElementById('div_anexos_lista_archivos');
                                if (!div) return [];
                                const html = div.innerHTML;
                                const results = [];
                                const fnName = "anexos_descargar_archivo(";
                                let startIdx = 0;
                                while (true) {
                                    const pos = html.indexOf(fnName, startIdx);
                                    if (pos === -1) break;
                                    startIdx = pos + fnName.length;
                                    const closePos = html.indexOf(')', startIdx);
                                    if (closePos === -1) break;
                                    const argsStr = html.substring(startIdx, closePos);
                                    const args = argsStr.split(',').map(a =>
                                        a.trim().replace(/^['"]/, '').replace(/['"]$/, '').trim()
                                    );
                                    if (args.length >= 3) {
                                        results.push({
                                            radicado: args[0],
                                            anex_codigo: args[1],
                                            arch_tipo: args[2],
                                            tipo_descarga: args.length >= 4 ? args[3] : 'download'
                                        });
                                    }
                                }
                                return results;
                            }
                        """)

                        seen = set()
                        unique_anexos = []
                        for a in anexo_data:
                            key = f"{a['radicado']}_{a['anex_codigo']}_{a['arch_tipo']}"
                            td = a.get('tipo_descarga', '').strip("'\"")
                            if key not in seen and 'embeded' not in td:
                                seen.add(key)
                                unique_anexos.append(a)

                        if unique_anexos:
                            os.makedirs(anexos_dir, exist_ok=True)
                            print(f"  Anexos: {len(unique_anexos)}")
                            for j, anexo in enumerate(unique_anexos):
                                try:
                                    rad = anexo['radicado']
                                    anex = anexo['anex_codigo']
                                    arch = anexo['arch_tipo']
                                    async with doc_page.expect_download(timeout=30000) as dl:
                                        await doc_page.evaluate(f"""
                                            () => {{
                                                anexos_descargar_archivo('{rad}', '{anex}', {arch});
                                            }}
                                        """)
                                    download = await dl.value
                                    fname = download.suggested_filename or f"anexo_{j + 1}.bin"
                                    await download.save_as(os.path.join(anexos_dir, fname))
                                    print(f"  ✓ Anexo: {fname}")
                                    await asyncio.sleep(1)
                                except Exception as e:
                                    print(f"  ⚠ Error anexo {j + 1}: {e}")
                        else:
                            print(f"  (Sin anexos)")

                except Exception as e:
                    import traceback
                    print(f"  ⚠ Error buscando anexos: {e}")
                    traceback.print_exc()

                await doc_page.close()

            except Exception as e:
                errores += 1
                print(f"  ✗ Error general: {e}")
                while len(context.pages) > 1:
                    try:
                        await context.pages[-1].close()
                    except Exception:
                        break

            # Datos para Excel
            num_anexos_descargados = 0
            anexos_path_check = os.path.join(doc_dir, "anexos")
            if os.path.isdir(anexos_path_check):
                num_anexos_descargados = len(os.listdir(anexos_path_check))

            all_docs_data.append({
                "No": total_docs,
                "De": doc.get("remitente", ""),
                "Asunto": doc.get("asunto", ""),
                "Fecha Documento": doc.get("fecha", ""),
                "Número Documento": numDoc,
                "No. Referencia": noReferencia,
                "Usuario Anterior": usuarioAnterior,
                "Carpeta": doc_folder,
                "Anexos": num_anexos_descargados,
            })
            print()

    # Generar Excel
    if all_docs_data:
        xlsx_name = f"quipux_{bandeja_nombre.lower().replace(' ', '_')}.xlsx"
        xlsx_path = os.path.join(download_dir, xlsx_name)
        generar_excel(all_docs_data, xlsx_path, bandeja_nombre)

    # Resumen
    print("\n" + "=" * 60)
    print(f"  RESUMEN — {bandeja_nombre}")
    print(f"  Total páginas:        {total_pages}")
    print(f"  Total documentos:     {total_docs}")
    if fecha_desde and fecha_hasta:
        print(f"  Rango de fechas:      {fecha_desde.strftime('%d/%m/%Y')} → {fecha_hasta.strftime('%d/%m/%Y')}")
        print(f"  Omitidos (fecha):     {omitidos_fecha}")
    print(f"  Descargados:          {descargados}")
    print(f"  Errores:              {errores}")
    print(f"  Carpeta:              {download_dir}")
    print("=" * 60)

    # Listar carpetas
    folders = sorted([
        d for d in os.listdir(download_dir)
        if os.path.isdir(os.path.join(download_dir, d)) and d.startswith("DOC_")
    ])
    if folders:
        print(f"\n  📁 Carpetas ({len(folders)}):")
        for folder in folders:
            full_path = os.path.join(download_dir, folder)
            main_files = [f for f in os.listdir(full_path) if os.path.isfile(os.path.join(full_path, f))]
            print(f"    📁 {folder}/ ({len(main_files)} doc)")
            anexos_path = os.path.join(full_path, "anexos")
            if os.path.isdir(anexos_path):
                anexo_files = os.listdir(anexos_path)
                if anexo_files:
                    print(f"       📁 anexos/ ({len(anexo_files)} archivos)")

    return total_docs, descargados, errores


# ─────────────────────────────────────────────────────────
# Programa principal con menú interactivo
# ─────────────────────────────────────────────────────────
async def run():
    async with async_playwright() as p:
        BASE_DIR = os.path.abspath(os.getenv("DOWNLOAD_DIR", "./mis_respaldos"))
        os.makedirs(BASE_DIR, exist_ok=True)

        print("=" * 60)
        print("  QUIPUX - Bot de descarga de documentos")
        print("=" * 60)

        browser = await p.chromium.launch(
            headless=False,
            args=["--ignore-certificate-errors"],
        )
        context = await browser.new_context(
            accept_downloads=True,
            ignore_https_errors=True,
        )
        await context.add_init_script("""
            Object.defineProperty(window, 'menubar', {
                get: () => ({ visible: false })
            });
            Object.defineProperty(window, 'toolbar', {
                get: () => ({ visible: false })
            });
        """)

        page = await context.new_page()
        page.set_default_timeout(60000)

        print("\n[1/3] Abriendo login.php...")
        await page.goto("https://quipux.espe.edu.ec/login.php", wait_until="domcontentloaded")
        await asyncio.sleep(2)
        print("       Ingresa tu usuario y clave en el navegador.")
        input("\n>>> Presiona ENTER cuando veas la bandeja de documentos... ")

        # ─── Encontrar página y frame ───
        print("\n[2/3] Buscando la bandeja de documentos...")
        target_page = None
        for pg in context.pages:
            try:
                url = pg.url
                if "index_frames" in url or "cuerpo" in url:
                    target_page = pg
            except Exception:
                pass
        if not target_page:
            for pg in context.pages:
                try:
                    _ = pg.url
                    target_page = pg
                except Exception:
                    pass
        if not target_page:
            print("ERROR: No se encontró ventana activa.")
            await browser.close()
            sys.exit(1)
        print(f"       ✓ Ventana: {target_page.url}")

        print("\n[3/3] Buscando frame principal...")
        await asyncio.sleep(2)

        # ─── Función para buscar mainFrame ───
        async def obtener_main_frame():
            for f in target_page.frames:
                if f.name == "mainFrame":
                    return f
            for f in target_page.frames:
                if f.url and "about:blank" not in f.url and f != target_page.main_frame:
                    try:
                        fc = await f.content()
                        if "mostrar_documento" in fc or "div_cuerpo" in fc:
                            return f
                    except Exception:
                        pass
            content = await target_page.content()
            if "mostrar_documento" in content or "div_cuerpo" in content:
                return target_page.main_frame
            return None

        # ─── Función para leer usuario actual ───
        async def obtener_usuario_actual():
            """Lee el nombre del usuario actual buscando en todos los frames."""
            nombre = None
            for frame in target_page.frames:
                try:
                    nombre = await frame.evaluate("""
                        () => {
                            const tds = document.querySelectorAll('td');
                            for (const td of tds) {
                                const text = td.innerText || '';
                                if (text.includes('Usuario actual:')) {
                                    const parts = text.split('Usuario actual:');
                                    if (parts[1]) return parts[1].trim().split('\\n')[0].trim();
                                }
                            }
                            const body = document.body?.innerText || '';
                            const m = body.match(/Usuario actual:\\s*(.+)/);
                            if (m) return m[1].trim().split('\\n')[0].trim();
                            return null;
                        }
                    """)
                    if nombre:
                        break
                except Exception:
                    continue
            return nombre or "Usuario desconocido"

        # ══════════════════════════════════════════════
        # SELECCIÓN INICIAL DE USUARIO
        # ══════════════════════════════════════════════
        print("\n" + "═" * 60)
        print("  Si necesitas cambiar de usuario en Quipux,")
        print("  hazlo ahora en el navegador.")
        print("═" * 60)
        input("\n>>> Presiona ENTER cuando estés listo con el usuario deseado... ")

        usuario_actual = await obtener_usuario_actual()

        # ── Nombre de carpeta para este usuario ──
        safe_user = re.sub(r'[^\w]', '_', usuario_actual)[:30]
        print(f"\n{'─' * 60}")
        print(f"  📁 Carpeta de descarga para este usuario")
        print(f"{'─' * 60}")
        print(f"  Base: {BASE_DIR}")
        print(f"  Carpeta por defecto: {safe_user}/")
        print(f"  (Dentro se crearán subcarpetas Recibidos/, Enviados/, Archivados/ y Carpetas Virtuales/)")
        print(f"\n  Ingresa el nombre de la carpeta o presiona ENTER")
        print(f"  para usar la carpeta por defecto.")
        nombre_carpeta = input(f"\n  📁 Nombre de carpeta [{safe_user}]: ").strip()
        if nombre_carpeta:
            nombre_carpeta = re.sub(r'[^\w\-. ]', '_', nombre_carpeta)[:50]
        else:
            nombre_carpeta = safe_user
        print(f"  ✓ Carpeta: {os.path.join(BASE_DIR, nombre_carpeta)}/")
        print(f"     ├── Recibidos/")
        print(f"     ├── Enviados/")
        print(f"     ├── Archivados/")
        print(f"     └── Carpetas Virtuales/")

        # ══════════════════════════════════════════════
        # MENÚ INTERACTIVO
        # ══════════════════════════════════════════════
        while True:
            print("\n" + "═" * 60)
            print("  ╔══════════════════════════════════════════════════╗")
            print("  ║          QUIPUX — Menú Principal                ║")
            print("  ╚══════════════════════════════════════════════════╝")
            print(f"\n  👤 Usuario: {usuario_actual}")
            print(f"  📁 Carpeta: {nombre_carpeta}/")
            print("\n  ─── Bandejas ───")
            print("  [1] 📥 Descargar Recibidos")
            print("  [2] 📤 Descargar Enviados")
            print("  [3] 📦 Descargar Archivados  (Otras Bandejas)")
            print("  [4] 📂 Descargar de Carpetas Virtuales")
            print("\n  ─── Opciones ───")
            print("  [5] 👤 Cambiar de usuario")
            print("  [0] 🚪 Salir")
            print("═" * 60)

            opcion = input("\n  Selecciona una opción: ").strip()

            if opcion == "0":
                print("\n  👋 ¡Hasta luego!")
                break

            elif opcion in ("1", "2", "3", "4"):
                opciones_bandeja = {
                    "1": ("Recibidos", "2", "📥"),
                    "2": ("Enviados", "8", "📤"),
                    "3": ("Archivados", "5", "📦"),
                    "4": ("Carpetas Virtuales", "0", "📂"),
                }
                bandeja_nombre, carpeta_codigo, emoji = opciones_bandeja[opcion]

                fecha_desde = None
                fecha_hasta = None

                dl_dir = os.path.join(BASE_DIR, nombre_carpeta, bandeja_nombre)

                # Navegar directamente a la bandeja con URL absoluta
                print(f"  Navegando a {bandeja_nombre}...")
                main_frame = await navegar_bandeja(target_page, bandeja_nombre, carpeta_codigo, pagina=1)
                if not main_frame:
                    print(f"  ⚠ Error: No se pudo navegar a {bandeja_nombre}.")
                    continue

                # --- LÓGICA ESPECIAL PARA CARPETAS VIRTUALES ---
                if opcion == "4":
                    print("\n  Cargando Carpetas Virtuales...")
                    await asyncio.sleep(5)
                    
                    try:
                        years = await main_frame.evaluate("""
                            () => {
                                const tds = document.querySelectorAll('td');
                                const results = [];
                                for (let td of tds) {
                                    const text = (td.innerText || '').trim();
                                    if (/^20\\d{2}$/.test(text)) {
                                        results.push(text);
                                    }
                                }
                                return [...new Set(results)].sort((a,b)=>b-a);
                            }
                        """)
                    except Exception:
                        years = []
                    
                    if years:
                        print("\n  Años detectados:")
                        for i, y in enumerate(years, 1):
                            print(f"  [{i}] {y}")
                        print("  [0] Ingresar año manualmente")
                        
                        opt = input("\n  Selecciona el año: ").strip()
                        if opt != "0" and opt.isdigit() and 1 <= int(opt) <= len(years):
                            year_to_click = years[int(opt)-1]
                        else:
                            year_to_click = input("  Ingresa el año exacto: ").strip()
                    else:
                        year_to_click = input("\n  📅 Ingresa el Año (ej. 2026): ").strip()
                        
                    if year_to_click:
                        print(f"  👉 Expandiendo '{year_to_click}'...")
                        try:
                            # Intentar clic en el texto exacto
                            await main_frame.locator(f"text='{year_to_click}'").first.click(timeout=5000)
                        except Exception:
                            print("  ⚠ No se pudo hacer clic automáticamente, por favor expande manualmente.")
                    
                    await asyncio.sleep(2)
                    carpeta_final_nombre = "Varios"
                    
                    print("\n  🔍 Evaluando el árbol de carpetas virtuales...")
                    await asyncio.sleep(4)
                    
                    js_extractor = """
                        () => {
                            const rows = document.querySelectorAll('tr');
                            const options = [];
                            let currentCategory = "Carpetas Principales";
                            
                            rows.forEach((row, idx) => {
                                // Ignorar filas que sean contenedores de otras tablas completas (evita basura acumulada)
                                if (row.querySelector('table')) return;
                                
                                const style = window.getComputedStyle(row);
                                if (style.display === 'none' || style.visibility === 'hidden') return;
                                
                                const text = row.innerText || '';
                                if (!text.includes('Activo') && !text.includes('Seleccionar')) return;
                                
                                const cleanText = text.replace(/\\n/g, ' ').trim();
                                const hasSelect = text.includes('Seleccionar');
                                
                                let folderName = '';
                                const match = cleanText.match(/^(.*?)(?:\\s+Activo|\\s+Inactivo)(?:\\s+Seleccionar)?/i);
                                if (match && match[1]) {
                                    folderName = match[1].trim().replace(/^[-+]+/, '').trim();
                                } else {
                                    folderName = cleanText.split('Activo')[0].trim().replace(/^[-+]+/, '').trim();
                                }
                                
                                // Filtrar cabeceras y años
                                if (!folderName || /^20\\d{2}$/.test(folderName) || folderName.includes('Nombre de Carpeta')) return;
                                
                                if (!hasSelect) {
                                    // Es un nivel intermedio (Categoría / Padre)
                                    currentCategory = folderName;
                                } else {
                                    // Es una carpeta final seleccionable
                                    // Limpiamos por si queda "Seleccionar" en el nombre
                                    folderName = folderName.replace(/Seleccionar.*$/i, '').trim();
                                    options.push({
                                        id: idx,
                                        name: folderName,
                                        category: currentCategory
                                    });
                                }
                            });
                            return options;
                        }
                    """
                    
                    try:
                        opciones_seleccionar = await main_frame.evaluate(js_extractor)
                    except Exception as e:
                        opciones_seleccionar = []
                        print(f"  ⚠ Advertencia evaluando DOM: {e}")
                    
                    if not opciones_seleccionar:
                        print("  ⚠ No se encontraron opciones disponibles para 'Seleccionar'.")
                        print("  1. Navega en la página web y expende el árbol manualmente (haz clic en los iconos +).")
                        print("  2. Una vez que las carpetas deseadas estén visibles en tu navegador, presiona ENTER para escanear de nuevo.")
                        input("  >>> Presiona ENTER cuando hayas expandido las opciones... ")
                        try:
                            opciones_seleccionar = await main_frame.evaluate(js_extractor)
                        except Exception:
                            opciones_seleccionar = []
                            
                    carpeta_final_nombre = "Varios"
                    
                    if opciones_seleccionar:
                        print("\n" + "═" * 70)
                        print("  📂 CARPETAS DISPONIBLES PARA SELECCIONAR")
                        print("═" * 70)
                        
                        from collections import defaultdict
                        grupos = defaultdict(list)
                        for opt in opciones_seleccionar:
                            grupos[opt['category']].append(opt)
                            
                        # Asignar un número global
                        contador = 1
                        mapa_opciones = {}
                        
                        for cat, lista in grupos.items():
                            print(f"\n  📁 {cat.upper()}")
                            for opt in lista:
                                print(f"      [{contador:2d}] {opt['name']}")
                                mapa_opciones[str(contador)] = opt
                                contador += 1
                                
                        print("\n" + "─" * 70)
                        print("  [0] Omitir lista / Abortar y seleccionar manualmente en el navegador")
                        
                        sel_idx = input("\n  🔢 Elige el número de la carpeta a descargar: ").strip()
                        
                        if sel_idx in mapa_opciones:
                            elegida = mapa_opciones[sel_idx]
                            carpeta_final_nombre = re.sub(r'[^\w\-. ]', '_', elegida['name'])
                            print(f"  👉 Clickeando en 'Seleccionar' para: {elegida['name']} ...")
                            try:
                                # Usamos JS directo para forzar el clic en el índice exacto extraído para saltar bloqueos visuales de Playwright
                                row_idx = elegida['id']
                                click_script = f"""
                                    () => {{
                                        const rows = document.querySelectorAll('tr');
                                        if ({row_idx} < rows.length) {{
                                            const row = rows[{row_idx}];
                                            const links = Array.from(row.querySelectorAll('a'));
                                            const selectLink = links.find(el => el.innerText.includes('Seleccionar'));
                                            if (selectLink) {{
                                                selectLink.click();
                                                return true;
                                            }}
                                        }}
                                        return false;
                                    }}
                                """
                                success = await main_frame.evaluate(click_script)
                                if success:
                                    print("  ✓ Carpeta virtual seleccionada con éxito (via JS).")
                                    await asyncio.sleep(5)
                                else:
                                    raise Exception("No se halló el link en el DOM o cambió de posición.")
                            except Exception as e:
                                print(f"  ⚠ Error al hacer clic automáticamente: {e}")
                                print("  Por favor haz clic manualmente en el navegador.")
                                input("  >>> Presiona ENTER cuando hayas seleccionado... ")
                        else:
                            print("  Selección omitida o no válida.")
                            input("  >>> Presiona ENTER cuando hayas hecho clic en 'Seleccionar' en el navegador web... ")
                    else:
                        print("\n  ⚠ No se detectó ninguna lista de carpetas.")
                        print("  Navega manualmente en la ventana del navegador y da clic en 'Seleccionar'.")
                        input("  >>> Presiona ENTER cuando veas los documentos de la carpeta listos para extraer... ")
                            
                    if year_to_click:
                        dl_dir = os.path.join(dl_dir, f"{year_to_click}_{carpeta_final_nombre}")
                # --- FIN LÓGICA ESPECIAL CARPETAS VIRTUALES ---

                if opcion == "4":
                    print("\n" + "═" * 70)
                    print("  ⚙  FILTRADO POR FECHAS EN CARPETAS VIRTUALES")
                    print("═" * 70)
                    print("  Nota: Las carpetas virtuales tienen su propio filtro nativo en Quipux.")
                    print("  En tu navegador verás que se muestran los documentos según el")
                    print("  rango de fechas seleccionado (ej. los últimos 3 meses).")
                    print("  ")
                    print("  Si deseas descargar MÁS O MENOS documentos, modifica las fechas")
                    print("  directamente en los campos 'Fecha Desde' y 'Fecha Hasta', y")
                    print("  haz clic en 'Buscar Documentos' en la pantalla.")
                    print("─" * 70)
                    input("  >>> Configura el filtro en tu navegador y presiona ENTER para extraer... ")
                    
                    fecha_desde = None
                    fecha_hasta = None
                    opcion_rango = "1"
                else:
                    # ── Sub-menú: rango de descarga ──
                    print(f"\n{'─' * 60}")
                    print(f"  {emoji} {bandeja_nombre.upper()} — Rango de descarga")
                    print(f"{'─' * 60}")
                    print("  [1] 📋 Descargar TODO")
                    print("  [2] 📅 Descargar por rango de fechas")
                    print("  [0] ↩  Volver al menú")
                    print(f"{'─' * 60}")

                    opcion_rango = input("\n  Selecciona una opción: ").strip()

                    if opcion_rango == "0":
                        continue

                    if opcion_rango == "2":
                        print("\n  Ingresa el rango de fechas para descargar documentos.")
                        print("  Solo se descargarán documentos cuya fecha esté dentro")
                        print("  del rango indicado (inclusive).")
                        print("  Formato: DD/MM/AAAA  (ejemplo: 01/01/2025)")

                        # ── Fecha DESDE ──
                        while True:
                            fecha_desde_input = input("\n  📅 Fecha DESDE: ").strip()
                            fecha_desde = parsear_fecha(fecha_desde_input)
                            if fecha_desde:
                                break
                            print("  ⚠ Formato de fecha no válido. Usa DD/MM/AAAA")
                            print("    Ejemplo: 01/01/2025")

                        # ── Fecha HASTA ──
                        while True:
                            fecha_hasta_input = input("  📅 Fecha HASTA: ").strip()
                            fecha_hasta = parsear_fecha(fecha_hasta_input)
                            if not fecha_hasta:
                                print("  ⚠ Formato de fecha no válido. Usa DD/MM/AAAA")
                                print("    Ejemplo: 31/12/2025")
                                continue
                            if fecha_hasta < fecha_desde:
                                print(f"  ⚠ La fecha HASTA ({fecha_hasta.strftime('%d/%m/%Y')}) no puede ser")
                                print(f"    anterior a la fecha DESDE ({fecha_desde.strftime('%d/%m/%Y')}).")
                                print(f"    Ingresa una fecha igual o posterior.")
                                continue
                            break

                        print(f"\n  ✔ Rango seleccionado: {fecha_desde.strftime('%d/%m/%Y')} → {fecha_hasta.strftime('%d/%m/%Y')}")
                    elif opcion_rango != "1":
                        print("  ⚠ Opción no válida")
                        continue

                print(f"\n{'=' * 60}")
                print(f"  {emoji} DESCARGA DE {bandeja_nombre.upper()}")
                print(f"  👤 {usuario_actual}")
                print(f"  📁 {dl_dir}")
                if fecha_desde and fecha_hasta:
                    print(f"  📅 Desde: {fecha_desde.strftime('%d/%m/%Y')}  Hasta: {fecha_hasta.strftime('%d/%m/%Y')}")
                else:
                    print(f"  📅 Descargando TODO")
                print(f"{'=' * 60}")

                await descargar_bandeja(target_page, main_frame, context, dl_dir, f"{bandeja_nombre} — {usuario_actual}", carpeta_codigo=carpeta_codigo, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)

                print("\n" + "─" * 60)
                input("  Presiona ENTER para volver al menú... ")

            elif opcion == "5":
                # ── Cambiar usuario ──
                print("\n" + "─" * 60)
                print("  👤 CAMBIAR DE USUARIO")
                print("─" * 60)
                print("\n  Ve al navegador y selecciona el usuario deseado.")
                print("  (Puedes hacerlo desde el panel izquierdo de Quipux)")
                input("\n>>> Presiona ENTER cuando hayas cambiado de usuario... ")

                await asyncio.sleep(2)
                usuario_actual = await obtener_usuario_actual()
                print(f"\n  ✓ Usuario detectado: {usuario_actual}")

                # ── Nombre de carpeta para el nuevo usuario ──
                safe_user = re.sub(r'[^\w]', '_', usuario_actual)[:30]
                print(f"\n{'─' * 60}")
                print(f"  📁 Carpeta de descarga para este usuario")
                print(f"{'─' * 60}")
                print(f"  Base: {BASE_DIR}")
                print(f"  Carpeta por defecto: {safe_user}/")
                print(f"  (Dentro se crearán subcarpetas Recibidos/, Enviados/, Archivados/ y Carpetas Virtuales/)")
                print(f"\n  Ingresa el nombre de la carpeta o presiona ENTER")
                print(f"  para usar la carpeta por defecto.")
                nombre_carpeta = input(f"\n  📁 Nombre de carpeta [{safe_user}]: ").strip()
                if nombre_carpeta:
                    nombre_carpeta = re.sub(r'[^\w\-. ]', '_', nombre_carpeta)[:50]
                else:
                    nombre_carpeta = safe_user
                print(f"  ✓ Carpeta: {os.path.join(BASE_DIR, nombre_carpeta)}/")
                print(f"     ├── Recibidos/")
                print(f"     ├── Enviados/")
                print(f"     ├── Archivados/")
                print(f"     └── Carpetas Virtuales/")

            else:
                print("  ⚠ Opción no válida")

        await browser.close()


asyncio.run(run())